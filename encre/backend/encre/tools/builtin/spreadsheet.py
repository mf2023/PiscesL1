#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Copyright © 2025-2026 Wenze Wei. All Rights Reserved.
#
# This file is part of Encre.
# The Encre project belongs to the Dunimd Team.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# You may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# DISCLAIMER: Users must comply with applicable AI regulations.
# Non-compliance may result in service termination or legal liability.

from __future__ import annotations
import csv
import io
import json
import os
from typing import Any

from encre.tools.base import build_tool


async def _spreadsheet_execute(**kwargs: Any) -> str:
    action = kwargs.get("action", "read")
    file_path = kwargs.get("file_path", "")
    sheet_name = kwargs.get("sheet_name", "")
    data = kwargs.get("data", "")
    _range = kwargs.get("range", "")

    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext == ".csv":
            return await _handle_csv(action, file_path, data)
        elif ext in (".xlsx", ".xls"):
            return await _handle_excel(action, file_path, sheet_name, data, _range)
        else:
            return f"Error: Unsupported file format: {ext}. Supported: .csv, .xlsx, .xls"
    except FileNotFoundError:
        return f"Error: File not found: {file_path}"
    except Exception as e:
        return f"Error processing spreadsheet: {e}"


async def _handle_csv(action: str, file_path: str, data: str) -> str:
    if action == "read":
        with open(file_path, "r", newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
        header = rows[0] if rows else []
        data_rows = rows[1:] if len(rows) > 1 else []
        return f"Header: {header}\nRows: {len(data_rows)}\n\n" + "\n".join(
            " | ".join(f"{c}" for c in row) for row in data_rows[:100]
        ) + (f"\n... ({len(data_rows) - 100} more rows)" if len(data_rows) > 100 else "")

    elif action == "write":
        if not data:
            return "Error: No data provided for write"
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            f.write(data)
        return f"Written CSV to {file_path}"

    elif action == "list_sheets":
        return "CSV files have a single sheet only"


async def _handle_excel(action: str, file_path: str, sheet_name: str, data: str, _range: str) -> str:
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl not installed. Install with: pip install openpyxl"

    if action == "read":
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb[sheet_name] if sheet_name else wb.active
        rows_data = []
        for row in sheet.iter_rows(values_only=True):
            rows_data.append([str(c) if c is not None else "" for c in row])
        header = rows_data[0] if rows_data else []
        data_rows = rows_data[1:] if len(rows_data) > 1 else []
        output = f"Sheet: {sheet.title}\nHeader: {header}\nRows: {len(data_rows)}\n\n"
        output += "\n".join(" | ".join(row) for row in data_rows[:100])
        if len(data_rows) > 100:
            output += f"\n... ({len(data_rows) - 100} more rows)"
        wb.close()
        return output

    elif action == "write":
        try:
            new_data = json.loads(data)
        except json.JSONDecodeError:
            return "Error: data must be valid JSON (array of arrays)"
        wb = openpyxl.Workbook()
        if sheet_name:
            wb.active.title = sheet_name
        sheet = wb.active
        for row in new_data:
            sheet.append(row)
        wb.save(file_path)
        return f"Written {len(new_data)} rows to {file_path}"

    elif action == "list_sheets":
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return f"Sheets: {json.dumps(sheets)}"


EncreSpreadsheetTool = build_tool(
    name="spreadsheet",
    description="Read, write, and manipulate CSV and Excel spreadsheets",
    input_schema={
        "type": "object",
        "properties": {
            "action": {
                "type": "string",
                "enum": ["read", "write", "list_sheets"],
                "description": "Action to perform",
            },
            "file_path": {
                "type": "string",
                "description": "Path to the spreadsheet file",
            },
            "sheet_name": {
                "type": "string",
                "description": "Sheet name (for Excel files)",
            },
            "data": {
                "type": "string",
                "description": "CSV or JSON data to write (for write action)",
            },
            "range": {
                "type": "string",
                "description": "Cell range to read (e.g. 'A1:C10', Excel only)",
            },
        },
        "required": ["action", "file_path"],
    },
    execute=_spreadsheet_execute,
    intents=["data", "research"],
    is_concurrency_safe=lambda _: True,
)
